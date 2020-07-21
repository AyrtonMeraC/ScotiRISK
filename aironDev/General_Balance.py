import numpy as np
import openpyxl
import pandas as pd
from pandas import ExcelWriter
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from aironDev import General_Register
excel='aironDev/Catalogo.xlsx'

class Data:
    def __init__(self):
        self.info=[]

class Reporte:
    def __init__(self,banco,codId,nombre,anho,mes):
        self.banco=banco
        self.codID=codId
        self.nombre=nombre
        self.anho=anho
        self.mes=mes
        self.registro=[]
        
class Catalogo:
    def __init__(self):
        self.cod=[]
        self.nombre=[]

def Tabla83_Disp():
    Tabla83=General_Register.Carga_Excel('aironDev/Tabla83.xlsx')
    list_data=[]
    for i in Tabla83.nombre:
        data={
            'Nombre':i,
            'Saldo':0,
        }
        list_data.append(data)
    hash = {k:v for k, v in zip(Tabla83.cod, list_data)}
    return hash

def Tabla73_Disp():
    Tabla73=General_Register.Carga_Excel('aironDev/Tabla73.xlsx')
    list_data=[]
    for i in Tabla73.nombre:
        data={
            'Nombre':i,
            'Saldo':0,
        }
        list_data.append(data)
    hash = {k:v for k, v in zip(Tabla73.cod, list_data)}
    return hash

def Carga_Balance(list_B,list_M,list_MC49):
    list_83=Tabla83_Disp()
    list_73=Tabla73_Disp()
    list_saldos=[]
    list_saldos49=[]
    list_Problemas=[]
    for i in list_M:
        if(i in list_B):

            if(list_M[i]['cOriFluC46Min']=='001' or list_M[i]['cOriFluC46May']=='002'):
                list_83[1]['Saldo']=list_83[1]['Saldo']+list_B[i]
            elif(list_M[i]['cOriFluC46Min']=='015' or list_M[i]['cOriFluC46May']=='016'):
                list_83[15]['Saldo']=list_83[15]['Saldo']+list_B[i]
            elif(list_M[i]['cOriFluC46Min']=='301' or list_M[i]['cOriFluC46May']=='302'):
                list_83[301]['Saldo']=list_83[301]['Saldo']+list_B[i]
            else:
                if(list_M[i]['cOriFluC46Min']!=''):
                    try:
                        list_83[int(list_M[i]['cOriFluC46Min'])]['Saldo']=list_83[int(list_M[i]['cOriFluC46Min'])]['Saldo']+list_B[i]
                    except:
                        data={
                            'Codigo':list_M[i]['cOriFluC46Min'],
                            'Cuenta':i,
                            'Saldo':list_B[i],
                        }
                        list_Problemas.append(data)
    for i in list_83:
        list_saldos.append(list_83[i]['Saldo'])
    
    for i in list_MC49:
        if(i in list_B):

            try:
                list_73[int(list_MC49[i]['Cod_Agr'])]['Saldo']=list_73[int(list_MC49[i]['Cod_Agr'])]['Saldo']+list_B[i]
            except:
                data={
                    'Codigo':list_MC49[i]['Cod_Agr'],
                    'Cuenta':i,
                    'Saldo':list_B[i],
                }
                list_Problemas.append(data)
    
    for i in list_73:
        list_saldos49.append(list_73[i]['Saldo'])

    return list_saldos,list_Problemas,list_saldos49

def Carga_M(excel):
    doc = openpyxl.load_workbook(excel)
    doc.get_sheet_names()
    hoja = doc.get_sheet_by_name('C46')
    hoja.rows
    list_Cu=[]
    list_CM=[]
    for filas in hoja.rows:
        data={
            'cOriFluC46Min':filas[4].value,
            'cOriFluC46May':filas[5].value,
        }
        list_Cu.append(str(filas[0].value))
        list_CM.append(data)
    hash = {k:v for k, v in zip(list_Cu[1:],list_CM[1:])}
    return hash

def Carga_MC49(excel):
    doc = openpyxl.load_workbook(excel)
    doc.get_sheet_names()
    hoja = doc.get_sheet_by_name('C49')
    hoja.rows
    list_Cu=[]
    list_CM=[]
    for filas in hoja.rows:
        codigo=filas[7].value
        data={
            'Cod_Agr':codigo[1:],
        }
        list_Cu.append(str(filas[9].value))
        list_CM.append(data)
    hash = {k:v for k, v in zip(list_Cu[1:],list_CM[1:])}
    return hash

def Carga_B(excel):
    doc = openpyxl.load_workbook(excel)
    doc.get_sheet_names()
    hoja = doc.get_sheet_by_name('SALDOS DIARIOS')
    hoja.rows
    list_B=[]
    list_C=[]
    list_S=[]
    flag=0
    count=0
    for filas in hoja.rows:
        if(filas[6].value!=None):
            largo=(len(filas[6].value)-1)
            key=str(filas[6].value[:largo])
            saldo=filas[10].value
            if(saldo==None):
                saldo=0

            if(len(list_B)>0):
                flag=0
                count=0
                for item in list_B:
                    if item[0] == key:
                        y = list(item)
                        y[1]=y[1]+saldo
                        x = tuple(y)
                        item=x
                        flag=1
                    list_B[count]=item
                    count=count+1
                if(flag==0):
                    list_B.append((key, saldo))
            else:
                list_B.append((key, saldo))
            
    for i in list_B[1:]:
        list_C.append(i[0])
        list_S.append(i[1])
        
    hash = {k:v for k, v in zip(list_C,list_S)}
    return hash




